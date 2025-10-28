#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de calendrier Excel avec mise en forme pour Fluxa
"""

import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_excel_calendar():
    """Crée un fichier Excel formaté pour le calendrier de publication"""

    # Charger le CSV
    csv_path = r'C:\Users\Utilisateur\Documents\Applications personnalisées\fluxa-artisans-automations-main\CALENDRIER_PUBLICATION_FLUXA.csv'

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f, delimiter=';')
        data = list(reader)

    # Créer le workbook
    wb = Workbook()

    # === FEUILLE 1 : CALENDRIER COMPLET ===
    ws1 = wb.active
    ws1.title = "Calendrier Publications"

    # En-têtes
    headers = ['Date', 'Jour', 'Horaire', 'Plateforme', 'Article', 'Titre',
               'Format', 'Priorité', 'Objectif SEO', 'Actions', 'Statut']

    # Couleurs par plateforme
    platform_colors = {
        'Blog Fluxa.fr': 'FF6B6B',  # Rouge
        'LinkedIn': '0077B5',  # Bleu LinkedIn
        'Facebook': '1877F2',  # Bleu Facebook
        'Instagram': 'E4405F',  # Rose Instagram
        'Instagram Stories': 'C13584'  # Violet Instagram
    }

    # Écrire les en-têtes
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(1, col, header)
        cell.font = Font(bold=True, color='FFFFFF', size=12)
        cell.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Écrire les données
    for row_idx, entry in enumerate(data, 2):
        ws1.cell(row_idx, 1, entry['Date'])
        ws1.cell(row_idx, 2, entry['Jour'])
        ws1.cell(row_idx, 3, entry['Horaire'])
        ws1.cell(row_idx, 4, entry['Plateforme'])
        ws1.cell(row_idx, 5, entry['Article'])
        ws1.cell(row_idx, 6, entry['Titre'])
        ws1.cell(row_idx, 7, entry['Format'])
        ws1.cell(row_idx, 8, entry['Priorite'])
        ws1.cell(row_idx, 9, entry['Objectif_SEO'])
        ws1.cell(row_idx, 10, entry['Actions'])
        ws1.cell(row_idx, 11, 'À FAIRE')  # Colonne statut

        # Couleur de ligne selon la plateforme
        platform = entry['Plateforme']
        color = platform_colors.get(platform, 'CCCCCC')

        for col in range(1, 12):
            cell = ws1.cell(row_idx, col)
            cell.fill = PatternFill(start_color=color + '20', end_color=color + '20', fill_type='solid')
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Bordures
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            cell.border = thin_border

        # Mise en gras pour priorité CRITIQUE
        if entry['Priorite'] == 'CRITIQUE':
            for col in range(1, 12):
                ws1.cell(row_idx, col).font = Font(bold=True)

    # Ajuster les largeurs de colonnes
    ws1.column_dimensions['A'].width = 12  # Date
    ws1.column_dimensions['B'].width = 12  # Jour
    ws1.column_dimensions['C'].width = 8   # Horaire
    ws1.column_dimensions['D'].width = 18  # Plateforme
    ws1.column_dimensions['E'].width = 35  # Article
    ws1.column_dimensions['F'].width = 45  # Titre
    ws1.column_dimensions['G'].width = 30  # Format
    ws1.column_dimensions['H'].width = 10  # Priorité
    ws1.column_dimensions['I'].width = 30  # Objectif SEO
    ws1.column_dimensions['J'].width = 50  # Actions
    ws1.column_dimensions['K'].width = 12  # Statut

    # Figer la première ligne
    ws1.freeze_panes = 'A2'

    # === FEUILLE 2 : VUE PAR SEMAINE ===
    ws2 = wb.create_sheet("Vue par Semaine")

    # Créer un résumé par semaine
    weeks = {}
    for entry in data:
        date = entry['Date']
        week = f"Semaine du {date}"

        if week not in weeks:
            weeks[week] = []

        weeks[week].append(entry)

    # En-têtes
    ws2.cell(1, 1, "Semaine").font = Font(bold=True, size=12)
    ws2.cell(1, 2, "Nb Publications").font = Font(bold=True, size=12)
    ws2.cell(1, 3, "Détails").font = Font(bold=True, size=12)

    row = 2
    for week, entries in list(weeks.items())[:26]:  # Limiter à 26 semaines
        ws2.cell(row, 1, week)
        ws2.cell(row, 2, len(entries))

        details = "\n".join([f"{e['Plateforme']}: {e['Article'][:30]}..." for e in entries])
        ws2.cell(row, 3, details)
        ws2.cell(row, 3).alignment = Alignment(wrap_text=True, vertical='top')

        row += 1

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 60

    # === FEUILLE 3 : GUIDE D'UTILISATION ===
    ws3 = wb.create_sheet("Guide Utilisation")

    guide_text = [
        ("GUIDE D'UTILISATION DU CALENDRIER FLUXA", 16, True),
        ("", 10, False),
        ("1. LÉGENDE DES COULEURS", 14, True),
        ("", 10, False),
        ("Rouge clair : Blog Fluxa.fr (PRIORITÉ ABSOLUE)", 11, False),
        ("Bleu LinkedIn : LinkedIn (Trafic B2B qualifié)", 11, False),
        ("Bleu Facebook : Facebook (Reach & Notoriété)", 11, False),
        ("Rose Instagram : Instagram Post (Engagement visuel)", 11, False),
        ("Violet : Instagram Stories (Rappels)", 11, False),
        ("", 10, False),
        ("2. PRIORITÉS", 14, True),
        ("", 10, False),
        ("🔴 CRITIQUE : Toujours publier en premier (Blog Fluxa.fr)", 11, False),
        ("🟠 HAUTE : Publication importante pour SEO", 11, False),
        ("🟡 MOYENNE : Publication complémentaire", 11, False),
        ("", 10, False),
        ("3. WORKFLOW DE PUBLICATION", 14, True),
        ("", 10, False),
        ("Étape 1 : Publier sur Blog Fluxa.fr (08h00)", 11, False),
        ("Étape 2 : Attendre 2h pour indexation Google", 11, False),
        ("Étape 3 : Publier sur LinkedIn (10h00)", 11, False),
        ("Étape 4 : Publier sur Facebook 2-3 jours après", 11, False),
        ("Étape 5 : Créer visuels Instagram puis publier", 11, False),
        ("", 10, False),
        ("4. ACTIONS PAR PLATEFORME", 14, True),
        ("", 10, False),
        ("Blog Fluxa.fr :", 12, True),
        ("- Vérifier URL optimisée (ex: /blog/automatisation-artisans-2025)", 10, False),
        ("- Ajouter images avec balises ALT", 10, False),
        ("- Liens internes vers 2-3 autres articles", 10, False),
        ("- Balises meta + Open Graph", 10, False),
        ("", 10, False),
        ("LinkedIn :", 12, True),
        ("- Extraire intro (2-3 paragraphes) + conclusion", 10, False),
        ("- Lien Fluxa.fr dans le texte du post", 10, False),
        ("- Hashtags: #Automatisation #TPE #Artisans #N8N", 10, False),
        ("- Répondre aux commentaires sous 2h", 10, False),
        ("", 10, False),
        ("Facebook :", 12, True),
        ("- Créer visuel accrocheur (Canva)", 10, False),
        ("- Texte 150-200 caractères", 10, False),
        ("- Lien blog en commentaire", 10, False),
        ("", 10, False),
        ("Instagram :", 12, True),
        ("- Carrousel 5-7 slides", 10, False),
        ("- Caption 200-300 caractères", 10, False),
        ("- 10-15 hashtags dont #FluxaAutomation", 10, False),
        ("- Lien dans bio", 10, False),
        ("", 10, False),
        ("5. SUIVI", 14, True),
        ("", 10, False),
        ("- Mettre à jour la colonne 'Statut' après chaque publication", 11, False),
        ("- Analyser Google Analytics chaque semaine", 11, False),
        ("- Ajuster horaires selon engagement", 11, False),
    ]

    for idx, (text, size, bold) in enumerate(guide_text, 1):
        cell = ws3.cell(idx, 1, text)
        cell.font = Font(size=size, bold=bold)
        cell.alignment = Alignment(wrap_text=True, vertical='top')

    ws3.column_dimensions['A'].width = 80

    # Sauvegarder
    output_path = r'C:\Users\Utilisateur\Documents\Applications personnalisées\fluxa-artisans-automations-main\CALENDRIER_PUBLICATION_FLUXA.xlsx'
    wb.save(output_path)

    return output_path

if __name__ == "__main__":
    print("Generation du calendrier Excel...")
    output = create_excel_calendar()
    print(f"OK - Calendrier Excel cree: {output}")
